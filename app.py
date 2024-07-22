from flask import Flask, render_template, jsonify
from openpyxl import load_workbook
import os

app = Flask(__name__)

def get_data_from_excel():
    # Carrega a planilha
    workbook = load_workbook('testebasopenxy.xlsm', keep_vba=True)
    sheet = workbook.active

    # Coleta dados
    data = []
    col_empty = [True] * (sheet.max_column + 1)  # Track empty columns

    # Iterar sobre as linhas e colunas
    for row in sheet.iter_rows(min_row=2, max_col=10, values_only=True):
        row_data = []
        for i, value in enumerate(row):
            if value is not None:
                col_empty[i] = False
            row_data.append(value)
        data.append(row_data)

    # Remover colunas vazias
    filtered_data = []
    for row in data:
        filtered_row = [value for i, value in enumerate(row) if not col_empty[i]]
        filtered_data.append(filtered_row)

    return filtered_data

@app.route("/")
def planilha():
    data = get_data_from_excel()
    return render_template('planilha.html', data=data)

@app.route("/api/data")
def get_data():
    data = get_data_from_excel()
    return jsonify(data)

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=True)
